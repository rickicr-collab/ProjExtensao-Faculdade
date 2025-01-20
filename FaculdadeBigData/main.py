import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import os
from openpyxl import Workbook, load_workbook

# Configurações da pasta para armazenamento dos arquivos
PASTA_INVENTARIO = "inventario"


# Função para verificar se a pasta existe e criá-la se necessário
def verificar_pasta():
    if not os.path.exists(PASTA_INVENTARIO):
        os.makedirs(PASTA_INVENTARIO)


# Função para criar ou carregar uma planilha
def carregar_planilha(mes: str, ano: str) -> Workbook:
    verificar_pasta()
    caminho = os.path.join(PASTA_INVENTARIO, f"{mes}_{ano}.xlsx")
    if os.path.exists(caminho):
        return load_workbook(caminho)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Nome", "Tipo", "Estoque Inicial", "Preço Unitário",
            "Quantidade Vendida", "Bebidas Faltantes ou Vencidas", "Mês", "Ano"
        ])
        wb.save(caminho)
        return wb


# Função para salvar dados em uma planilha
def salvar_dados(mes: str, ano: str, dados: list) -> None:
    wb = carregar_planilha(mes, ano)
    ws = wb.active
    for dado in dados:
        ws.append(dado)
    wb.save(os.path.join(PASTA_INVENTARIO, f"{mes}_{ano}.xlsx"))


# Função para atualizar os dados em uma planilha
def atualizar_dados(mes: str, ano: str, dados: list) -> None:
    caminho = os.path.join(PASTA_INVENTARIO, f"{mes}_{ano}.xlsx")
    wb = load_workbook(caminho)
    ws = wb.active
    for i, dado in enumerate(dados):
        for j, valor in enumerate(dado):
            ws.cell(row=i + 2, column=j + 1, value=valor)
    wb.save(caminho)


# Função para excluir dados da planilha
def excluir_dados(mes: str, ano: str, linha: int) -> None:
    caminho = os.path.join(PASTA_INVENTARIO, f"{mes}_{ano}.xlsx")
    wb = load_workbook(caminho)
    ws = wb.active
    ws.delete_rows(linha + 2)  # Ajuste para corresponder ao índice da planilha
    wb.save(caminho)


# Função para gerar análise de dados
def analisar_dados(mes: str, ano: str) -> None:
    caminho = os.path.join(PASTA_INVENTARIO, f"{mes}_{ano}.xlsx")
    if not os.path.exists(caminho):
        messagebox.showerror("Erro", "Nenhuma planilha encontrada para o mês e ano selecionados.")
        return

    df = pd.read_excel(caminho)
    valor_inicial = (df["Estoque Inicial"] * df["Preço Unitário"]).sum()
    lucro_final = (df["Quantidade Vendida"] * df["Preço Unitário"]).sum() - valor_inicial
    valores_perdidos = df["Bebidas Faltantes ou Vencidas"].dropna().apply(pd.to_numeric, errors='coerce').sum()
    total_perdido = valores_perdidos
    produto_mais_vendido = df.loc[df["Quantidade Vendida"].idxmax()]["Nome"]
    produto_menos_vendido = df.loc[df["Quantidade Vendida"].idxmin()]["Nome"]

    messagebox.showinfo("Análise de Dados",
                        f"Valor Inicial Investido: R$ {valor_inicial:.2f}\n"
                        f"Lucro Final: R$ {lucro_final:.2f}\n"
                        f"Total Perdido: R$ {total_perdido:.2f}\n"
                        f"Produto Mais Vendido: {produto_mais_vendido}\n"
                        f"Produto Menos Vendido: {produto_menos_vendido}")


# Função para atualizar a área de visualização
def atualizar_visualizacao() -> None:
    mes = combo_mes.get()
    ano = entrada_ano.get()
    caminho = os.path.join(PASTA_INVENTARIO, f"{mes}_{ano}.xlsx")
    if os.path.exists(caminho):
        df = pd.read_excel(caminho)
        for i in treeview.get_children():
            treeview.delete(i)
        for _, row in df.iterrows():
            treeview.insert("", "end", values=list(row))
    else:
        messagebox.showerror("Erro", "Nenhuma planilha encontrada para o mês e ano selecionados.")


# Função para adicionar dados
def adicionar_dados() -> None:
    nome = entrada_nome.get()
    tipo = entrada_tipo.get()
    estoque_inicial = entrada_estoque_inicial.get()
    preco_unitario = entrada_preco_unitario.get()
    quantidade_vendida = entrada_quantidade_vendida.get()
    bebidas_faltantes = entrada_bebidas_faltantes.get()
    mes = combo_mes.get()
    ano = entrada_ano.get()

    dados = [[nome, tipo, estoque_inicial, preco_unitario, quantidade_vendida, bebidas_faltantes, mes, ano]]
    salvar_dados(mes, ano, dados)
    atualizar_visualizacao()
    limpar_campos()


# Função para editar dados
def editar_dados() -> None:
    mes = combo_mes.get()
    ano = entrada_ano.get()
    caminho = os.path.join(PASTA_INVENTARIO, f"{mes}_{ano}.xlsx")
    if not os.path.exists(caminho):
        messagebox.showerror("Erro", "Nenhuma planilha encontrada para o mês e ano selecionados.")
        return

    selecionado = treeview.selection()
    if not selecionado:
        messagebox.showerror("Erro", "Nenhum item selecionado.")
        return

    dados = [treeview.item(item, 'values') for item in selecionado]
    for dado in dados:
        dado[0] = entrada_nome.get()
        dado[1] = entrada_tipo.get()
        dado[2] = entrada_estoque_inicial.get()
        dado[3] = entrada_preco_unitario.get()
        dado[4] = entrada_quantidade_vendida.get()
        dado[5] = entrada_bebidas_faltantes.get()
    atualizar_dados(mes, ano, dados)
    atualizar_visualizacao()
    limpar_campos()


# Função para excluir dados
def excluir_dados_funcao() -> None:
    mes = combo_mes.get()
    ano = entrada_ano.get()
    caminho = os.path.join(PASTA_INVENTARIO, f"{mes}_{ano}.xlsx")
    if not os.path.exists(caminho):
        messagebox.showerror("Erro", "Nenhuma planilha encontrada para o mês e ano selecionados.")
        return

    selecionado = treeview.selection()
    if not selecionado:
        messagebox.showerror("Erro", "Nenhum item selecionado.")
        return

    for item in selecionado:
        linha = treeview.index(item)
        excluir_dados(mes, ano, linha)
    atualizar_visualizacao()


# Função para limpar os campos de inserção
def limpar_campos() -> None:
    entrada_nome.delete(0, tk.END)
    entrada_tipo.delete(0, tk.END)
    entrada_estoque_inicial.delete(0, tk.END)
    entrada_preco_unitario.delete(0, tk.END)
    entrada_quantidade_vendida.delete(0, tk.END)
    entrada_bebidas_faltantes.delete(0, tk.END)
    combo_mes.set("Selecione o mês")
    entrada_ano.delete(0, tk.END)


# Função para abrir a janela de análise de dados
def abrir_janela_analise() -> None:
    janela_analise = tk.Toplevel(root)
    janela_analise.title("Análise de Dados")

    tk.Label(janela_analise, text="Mês").grid(row=0, column=0, padx=5, pady=5)
    tk.Label(janela_analise, text="Ano").grid(row=1, column=0, padx=5, pady=5)

    mes_analise = ttk.Combobox(janela_analise, values=[
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ])
    mes_analise.grid(row=0, column=1, padx=5, pady=5)

    ano_analise = tk.Entry(janela_analise)
    ano_analise.grid(row=1, column=1, padx=5, pady=5)

    tk.Button(janela_analise, text="Analisar",
              command=lambda: analisar_dados(mes_analise.get(), ano_analise.get())).grid(row=2, column=1, padx=5,
                                                                                         pady=5)


root = tk.Tk()
root.title("Sistema de Estoque")

# Configuração da interface gráfica
tk.Label(root, text="Nome").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entrada_nome = tk.Entry(root)
entrada_nome.grid(row=0, column=1, padx=5, pady=5)

tk.Label(root, text="Tipo").grid(row=1, column=0, padx=5, pady=5, sticky="w")
entrada_tipo = tk.Entry(root)
entrada_tipo.grid(row=1, column=1, padx=5, pady=5)

tk.Label(root, text="Estoque Inicial").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entrada_estoque_inicial = tk.Entry(root)
entrada_estoque_inicial.grid(row=2, column=1, padx=5, pady=5)

tk.Label(root, text="Preço Unitário").grid(row=3, column=0, padx=5, pady=5, sticky="w")
entrada_preco_unitario = tk.Entry(root)
entrada_preco_unitario.grid(row=3, column=1, padx=5, pady=5)

tk.Label(root, text="Quantidade Vendida").grid(row=4, column=0, padx=5, pady=5, sticky="w")
entrada_quantidade_vendida = tk.Entry(root)
entrada_quantidade_vendida.grid(row=4, column=1, padx=5, pady=5)

tk.Label(root, text="Bebidas Faltantes ou Vencidas").grid(row=5, column=0, padx=5, pady=5, sticky="w")
entrada_bebidas_faltantes = tk.Entry(root)
entrada_bebidas_faltantes.grid(row=5, column=1, padx=5, pady=5)

tk.Label(root, text="Mês").grid(row=6, column=0, padx=5, pady=5, sticky="w")
combo_mes = ttk.Combobox(root, values=[
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
])
combo_mes.grid(row=6, column=1, padx=5, pady=5)
combo_mes.set("Selecione o mês")

tk.Label(root, text="Ano").grid(row=7, column=0, padx=5, pady=5, sticky="w")
entrada_ano = tk.Entry(root)
entrada_ano.grid(row=7, column=1, padx=5, pady=5)

tk.Button(root, text="Adicionar", command=adicionar_dados).grid(row=8, column=0, padx=5, pady=5)
tk.Button(root, text="Editar", command=editar_dados).grid(row=8, column=1, padx=5, pady=5)
tk.Button(root, text="Excluir", command=excluir_dados_funcao).grid(row=8, column=2, padx=5, pady=5)
tk.Button(root, text="Analisar", command=abrir_janela_analise).grid(row=8, column=3, padx=5, pady=5)

# Treeview para visualização dos dados
colunas = (
"Nome", "Tipo", "Estoque Inicial", "Preço Unitário", "Quantidade Vendida", "Bebidas Faltantes ou Vencidas", "Mês",
"Ano")
treeview = ttk.Treeview(root, columns=colunas, show="headings")
for col in colunas:
    treeview.heading(col, text=col)
    treeview.column(col, width=100)
treeview.grid(row=9, column=0, columnspan=4, padx=5, pady=5)

root.mainloop()
